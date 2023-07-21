import os
import sys
import traceback

import django
import openpyxl as openpyxl

sys.path[0] = '/app/'
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from apps.games.models import Game


def load_games():
    try:
        wb = openpyxl.load_workbook(
            filename=f"apps/games/excel/games.xlsx", data_only=True
        )
        sheet = wb[wb.sheetnames[0]]
        row_num, column_num = 2, 2

        next_game_name = sheet.cell(row=row_num, column=column_num).value
        while next_game_name is not None:
            game_name = next_game_name
            min_players = sheet.cell(row=row_num, column=column_num + 1).value
            max_players = sheet.cell(row=row_num, column=column_num + 2).value
            rec_play = sheet.cell(row=row_num, column=column_num + 3).value
            link = sheet.cell(row=row_num, column=column_num + 4).value
            description = sheet.cell(row=row_num, column=column_num + 5).value
            exp_time = sheet.cell(row=row_num, column=column_num + 6).value
            game, _ = Game.objects.get_or_create(name=game_name)
            game.name = game_name
            if min_players is not None:
                game.min_players = min_players
            if max_players is not None:
                game.max_players = max_players or 2

            if rec_play is not None:
                game.recommended_players = rec_play or 2
            print(game_name, link or "-")
            if link is not None:
                game.link = link

            if exp_time is not None:
                game.expected_length = exp_time or 2
            game.save()

            row_num += 1
            next_game_name = sheet.cell(row=row_num, column=column_num).value

    except Exception as e:
        traceback.print_exc()


if __name__ == '__main__':
    load_games()
